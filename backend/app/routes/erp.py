"""ERP and Sage integration routes with period filters, preview, audit log and export history."""
import io
import os
import uuid
from typing import Optional
from datetime import datetime, time

# FIX #5 — suppression du double import (Session, get_db apparaissaient deux fois)
from sqlalchemy.orm import Session, joinedload
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response

from ..database import get_db
from ..models import (
    Invoice,
    User,
    InvoiceStatus,
    SageExportBatch,
    SageExportJob,
    AuditLog,
    Notification,
    NotificationType,
)
from ..middleware.rbac import require_directeur
from ..services.sage_export import (
    SageExportConfig,
    build_sage_bytes,
    build_sage_zip_by_period,
    validate_sage_invoice,
    build_invoice_sage_lines,
)
from ..services.sage_auto_importer import sage_auto_importer

router = APIRouter(prefix="/erp", tags=["ERP / Sage"])


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def sage_config() -> SageExportConfig:
    return SageExportConfig(
        journal_code=os.getenv("SAGE_JOURNAL_CODE", "VTE"),
        client_account=os.getenv("SAGE_CLIENT_ACCOUNT", "34210000"),
        sales_account=os.getenv("SAGE_SALES_ACCOUNT", "71243000"),
        vat_account=os.getenv("SAGE_VAT_ACCOUNT", "44550000"),
        default_tiers_code=os.getenv("SAGE_DEFAULT_TIERS_CODE", "B01"),
    )


def parse_date(value: Optional[str]):
    if not value:
        return None
    return datetime.combine(datetime.fromisoformat(value).date(), time.min)


def invoice_query(
    db: Session,
    invoice_id: Optional[uuid.UUID] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    paid_only: bool = False,
    only_not_exported: bool = False,
):
    """Base query — charge toujours client, profil, contrat et items en une passe."""
    q = (
        db.query(Invoice)
        .options(
            joinedload(Invoice.client).joinedload(User.profile),
            joinedload(Invoice.contract),
            joinedload(Invoice.items),
        )
        .filter(
            Invoice.total.isnot(None),
            Invoice.total > 0,
        )
        .order_by(
            Invoice.issue_date.asc(),
            Invoice.invoice_number.asc(),
        )
    )

    if invoice_id:
        return q.filter(Invoice.id == invoice_id)

    if paid_only:
        q = q.filter(
            Invoice.status.in_([
                InvoiceStatus.PAID,
                InvoiceStatus.PENDING,
            ])
        )

    if only_not_exported:
        q = q.filter(Invoice.exported_to_sage == False)  # noqa: E712

    start_dt = parse_date(start_date)
    end_dt = parse_date(end_date)

    if start_dt:
        q = q.filter(Invoice.issue_date >= start_dt)

    if end_dt:
        q = q.filter(
            Invoice.issue_date <= end_dt.replace(hour=23, minute=59, second=59)
        )

    return q


def collect_sage_errors(invoices):
    errors = []
    for inv in invoices:
        errors.extend(validate_sage_invoice(inv))
    return errors


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/sage/mapping")
def sage_mapping(current_user: User = Depends(require_directeur)):
    cfg = sage_config()
    return {
        "sage_format_name": "import vente us",
        "type_fichier": "Délimité",
        "origine": "Windows",
        "delimiteur_enregistrement": "Retour chariot / CRLF",
        "delimiteur_champ": "Point-virgule ;",
        "entete": "Aucun",
        "format_date": "JJMMAA",
        "format_montant": "2 décimales, séparateur décimal Virgule, séparateur milliers Aucun",
        "mapping": [
            "1 Code journal",
            "2 Date pièce",
            "3 N° pièce",
            "4 N° facture",
            "5 Référence",
            "6 N° compte général",
            "7 N° compte tiers",
            "8 Libellé écriture",
            "9 Date échéance",
            "10 Débit",
            "11 Crédit",
        ],
        "config": cfg.__dict__,
        "sample": [
            "VTE;210326;26001;FAC-DEMO-2026-001;CTR-DEMO-2026-001;34210000;B01;DOM FACDEMO2026001 ATLAS DIGITAL SERVICES;200426;42000,00;",
            "VTE;210326;26001;FAC-DEMO-2026-001;CTR-DEMO-2026-001;71243000;;DOM FACDEMO2026001 ATLAS DIGITAL SERVICES;;;35000,00",
            "VTE;210326;26001;FAC-DEMO-2026-001;CTR-DEMO-2026-001;44550000;;TVA 20 DOM FACDEMO2026001 ATLAS DIGITAL SERVICES;;;7000,00",
        ],
    }


@router.get("/sage/preview")
def preview_sage_export(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    only_not_exported: bool = True,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    # FIX #6 — invoice_query charge déjà les items/client/contrat,
    # pas besoin de .options() supplémentaire ici
    invoices = (
        invoice_query(
            db,
            start_date=start_date,
            end_date=end_date,
            only_not_exported=only_not_exported,
        )
        .limit(20)
        .all()
    )
    errors = collect_sage_errors(invoices)
    preview = []
    for inv in invoices:
        preview.extend(build_invoice_sage_lines(inv, sage_config()))

    return {
        "errors": errors,
        "preview": preview[:60],
        "invoice_count_previewed": len(invoices),
        "total_amount": float(sum(inv.total or 0 for inv in invoices)),
        "filters": {
            "start_date": start_date,
            "end_date": end_date,
            "only_not_exported": only_not_exported,
        },
    }


@router.get("/sage/export/txt")
def export_sage_txt(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    only_not_exported: bool = True,
    mark_as_exported: bool = True,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    # FIX #7 — utilisait db.query(Invoice) directement, ignorait tous les filtres.
    # Maintenant passe par invoice_query comme les autres routes.
    invoices = invoice_query(
        db,
        start_date=start_date,
        end_date=end_date,
        only_not_exported=only_not_exported,
    ).all()

    if not invoices:
        raise HTTPException(status_code=400, detail="Aucune facture trouvée.")

    errors = collect_sage_errors(invoices)
    if errors:
        raise HTTPException(
            status_code=400,
            detail={"message": "Validation Sage échouée", "errors": errors},
        )

    content = build_sage_bytes(invoices, sage_config())
    filename = f"SAGE_EXPORT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    if mark_as_exported:
        for inv in invoices:
            inv.exported_to_sage = True
            inv.sage_exported_at = datetime.utcnow()
        db.commit()

    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/plain; charset=windows-1252",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/sage/export/drop-folder")
def export_sage_to_drop_folder(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    invoices = invoice_query(
        db, start_date=start_date, end_date=end_date, only_not_exported=True
    ).all()
    errors = collect_sage_errors(invoices)
    if errors:
        raise HTTPException(
            status_code=400,
            detail={"message": "Validation Sage échouée", "errors": errors},
        )

    drop_dir = os.getenv(
        "SAGE_AUTO_IMPORT_FOLDER",
        os.getenv("SAGE_EXPORT_DROP_DIR", r"C:\SAGE_AUTO_IMPORT\pending"),
    )
    os.makedirs(drop_dir, exist_ok=True)
    filename = (
        f"SAGE_VENTES_{start_date or 'DEBUT'}_{end_date or 'FIN'}_JJMMAA.txt"
        .replace("-", "")
    )
    path = os.path.abspath(os.path.join(drop_dir, filename))
    with open(path, "wb") as f:
        f.write(build_sage_bytes(invoices, sage_config()))

    db.add(AuditLog(
        actor_id=current_user.id,
        action="prepare_sage_drop_folder",
        entity_type="sage_export",
        description=f"Fichier Sage préparé dans le dossier: {path}",
        meta={"path": path, "invoice_count": len(invoices)},
    ))
    db.commit()

    return {
        "message": (
            "Fichier Sage préparé dans le dossier surveillé. "
            "Si SAGE_AUTO_IMPORT_ENABLED=true sur Windows, Sage va l'importer automatiquement."
        ),
        "path": path,
        "filename": filename,
        "date_format": "JJMMAA",
        "automation_note": (
            "Backend Windows + SAGE_AUTO_IMPORT_ENABLED=true = import automatique. "
            "Docker/Linux: génération TXT seulement."
        ),
        "auto_importer": sage_auto_importer.status(),
    }


@router.get("/sage/auto-import/status")
def sage_auto_import_status(current_user: User = Depends(require_directeur)):
    return sage_auto_importer.status()


@router.get("/sage/export/zip")
def export_sage_zip(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    invoices = invoice_query(
        db, start_date=start_date, end_date=end_date, only_not_exported=True
    ).all()
    errors = collect_sage_errors(invoices)
    if errors:
        raise HTTPException(
            status_code=400,
            detail={"message": "Validation Sage échouée", "errors": errors},
        )
    content = build_sage_zip_by_period(invoices, sage_config())
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=SAGE_VENTES_PAR_PERIODE.zip"},
    )


@router.get("/sage/invoices/{invoice_id}/txt")
def export_one_invoice_sage_txt(
    invoice_id: uuid.UUID,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    inv = invoice_query(db, invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    errors = validate_sage_invoice(inv)
    if errors:
        raise HTTPException(
            status_code=400,
            detail={"message": "Validation Sage échouée", "errors": errors},
        )
    content = build_sage_bytes([inv], sage_config())
    name = (inv.invoice_number or "FACTURE").replace("-", "").replace("/", "")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/plain; charset=windows-1252",
        headers={"Content-Disposition": f"attachment; filename=SAGE_{name}.txt"},
    )


@router.get("/sage/export/excel")
def export_sage_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    only_not_exported: bool = True,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    """Export Excel with the same columns as Sage accounting canvas."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    invoices = invoice_query(
        db,
        start_date=start_date,
        end_date=end_date,
        only_not_exported=only_not_exported,
    ).all()
    errors = collect_sage_errors(invoices)
    cfg = sage_config()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Canvas Comptabilite"
    headers = [
        "Code journal", "Date pièce", "N° pièce", "N° facture", "Référence",
        "N° compte général", "N° compte tiers", "Libellé écriture",
        "Date échéance", "Débit", "Crédit",
    ]
    ws.append(headers)

    for inv in invoices:
        for line in build_invoice_sage_lines(inv, cfg):
            ws.append(line.split(";"))

    ws_control = wb.create_sheet("Controle")
    ws_control.append(["Période", f"{start_date or 'Début'} → {end_date or 'Fin'}"])
    ws_control.append(["Factures", len(invoices)])
    ws_control.append(["Total TTC", float(sum(inv.total or 0 for inv in invoices))])
    ws_control.append(["Erreurs", len(errors)])
    ws_control.append([])
    ws_control.append([
        "N° facture", "Référence", "Client", "Email",
        "Date pièce", "Échéance", "HT", "TVA", "TTC", "Exportée", "Validation",
    ])
    for inv in invoices:
        profile = inv.client.profile if inv.client and inv.client.profile else None
        client_name = (
            profile.company_name
            if profile and profile.company_name
            else (
                profile.full_name
                if profile
                else inv.client.email if inv.client else "-"
            )
        )
        inv_errors = validate_sage_invoice(inv)
        reference = getattr(inv.contract, "contract_number", None) or inv.invoice_number
        ws_control.append([
            inv.invoice_number,
            reference,
            client_name,
            inv.client.email if inv.client else "-",
            inv.issue_date.strftime("%d/%m/%Y") if inv.issue_date else "",
            inv.due_date.strftime("%d/%m/%Y") if inv.due_date else "",
            float(inv.subtotal or 0),
            float(inv.tax_amount or 0),
            float(inv.total or 0),
            "Oui" if inv.exported_to_sage else "Non",
            "OK" if not inv_errors else " | ".join(inv_errors),
        ])

    ws_errors = wb.create_sheet("Erreurs")
    ws_errors.append(["Erreur"])
    if errors:
        for error in errors:
            ws_errors.append([error])
    else:
        ws_errors.append(["Aucune erreur détectée"])

    header_fill = PatternFill("solid", fgColor="4B5563")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    for sheet in wb.worksheets:
        header_row = 6 if sheet.title == "Controle" else 1
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
                cell.border = Border(bottom=thin)
        for cell in sheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2" if sheet.title != "Controle" else "A7"
        for col in range(1, sheet.max_column + 1):
            max_len = max(
                len(str(sheet.cell(row, col).value or ""))
                for row in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[get_column_letter(col)].width = min(
                max(12, max_len + 3), 36
            )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = (
        f"SAGE_CANVAS_COMPTABILITE_{start_date or 'DEBUT'}_{end_date or 'FIN'}.xlsx"
        .replace("-", "")
    )

    db.add(AuditLog(
        actor_id=current_user.id,
        action="export_sage_excel_canvas",
        entity_type="sage_export",
        description=f"Export Excel canvas comptabilité généré pour {len(invoices)} facture(s).",
        meta={
            "start_date": start_date,
            "end_date": end_date,
            "invoice_count": len(invoices),
            "errors": errors,
        },
    ))
    db.commit()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/sage/history")
def sage_export_history(
    limit: int = Query(25, ge=1, le=100),
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(SageExportBatch)
        .order_by(SageExportBatch.created_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": str(r.id),
            "filename": r.filename,
            "export_type": r.export_type,
            "status": r.status,
            "invoice_count": r.invoice_count,
            "total_amount": r.total_amount,
            "period_start": r.period_start,
            "period_end": r.period_end,
            "created_at": r.created_at,
            "errors": r.errors or [],
        }
        for r in rows
    ]


@router.post("/sage/prepare-export")
def prepare_sage_export(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    only_not_exported: bool = True,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db),
):
    # FIX #8 — invoice_query charge déjà items/client/contrat.
    # Suppression du .options() redondant qui causait un double chargement.
    invoices = invoice_query(
        db,
        start_date=start_date,
        end_date=end_date,
        only_not_exported=only_not_exported,
    ).all()

    if not invoices:
        raise HTTPException(
            status_code=400,
            detail="Aucune facture à exporter (toutes déjà exportées ou aucune trouvée).",
        )

    errors = collect_sage_errors(invoices)
    if errors:
        # FIX #9 — retourne les erreurs détaillées pour que le frontend puisse les afficher
        raise HTTPException(
            status_code=400,
            detail={"message": "Validation Sage échouée", "errors": errors},
        )

    # FIX #10 — build_sage_bytes peut lever ValueError si aucune ligne générée.
    # On l'attrape et on retourne un 400 clair au lieu d'un 500 générique.
    try:
        content_bytes = build_sage_bytes(invoices, sage_config())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    txt_content = content_bytes.decode("cp1252")
    filename = f"SAGE_EXPORT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    job = SageExportJob(
        filename=filename,
        content=txt_content,
        status="pending",
    )
    db.add(job)

    for inv in invoices:
        inv.exported_to_sage = True
        inv.sage_exported_at = datetime.utcnow()

    db.add(AuditLog(
        actor_id=current_user.id,
        action="prepare_sage_export",
        entity_type="sage_export",
        description=f"Export Sage préparé: {filename} ({len(invoices)} facture(s))",
        meta={
            "filename": filename,
            "invoice_count": len(invoices),
            "start_date": start_date,
            "end_date": end_date,
        },
    ))

    db.commit()
    db.refresh(job)

    return {
        "message": "Export Sage préparé",
        "filename": filename,
        "job_id": job.id,
        "invoice_count": len(invoices),
    }


@router.get("/sage/agent/pending")
def get_pending_exports(db: Session = Depends(get_db)):
    jobs = db.query(SageExportJob).filter(SageExportJob.status == "pending").all()
    return [{"id": j.id, "filename": j.filename} for j in jobs]


@router.get("/sage/agent/download/{job_id}")
def download_export(job_id: int, db: Session = Depends(get_db)):
    job = db.query(SageExportJob).filter(SageExportJob.id == job_id).first()

    if not job:
        raise HTTPException(status_code=404, detail="Export introuvable")

    job.status = "downloaded"
    job.downloaded_at = datetime.utcnow()
    db.commit()

    return Response(
        content=job.content,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={job.filename}"},
    )
