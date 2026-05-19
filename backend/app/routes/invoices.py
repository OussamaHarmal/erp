"""
Invoice Routes
Electronic invoicing with item management and Excel export
"""
import io
import uuid
import os
from pathlib import Path
from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..models import Invoice, InvoiceItem, User, UserRole, InvoiceStatus, Contract
from ..schemas.schemas import InvoiceCreate, InvoiceUpdate, InvoiceResponse
from ..middleware.rbac import require_directeur, require_any_authenticated, check_resource_access
from ..services.invoice_exports import generate_invoice_pdf
from ..services.email_service import send_email_with_attachments

router = APIRouter(prefix="/invoices", tags=["Invoices"])


def generate_invoice_number(db: Session) -> str:
    count = db.query(Invoice).count()
    year = datetime.now().year
    # Keep numbering aligned with the visual template and ERP references.
    return f"FACT-{year}-{str(count + 1).zfill(5)}"

def add_months(dt: datetime, months: int) -> datetime:
    from calendar import monthrange
    month = dt.month - 1 + int(months)
    year = dt.year + month // 12
    month = month % 12 + 1
    day = min(dt.day, monthrange(year, month)[1])
    return dt.replace(year=year, month=month, day=day)


def get_client_email(invoice: Invoice) -> str:
    profile = invoice.client.profile if invoice.client and invoice.client.profile else None
    return getattr(profile, "company_email", None) or getattr(invoice.client, "email", "")



@router.get("", response_model=List[InvoiceResponse])
def list_invoices(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=200),
    status: Optional[str] = None,
    current_user: User = Depends(require_any_authenticated),
    db: Session = Depends(get_db)
):
    """
    List invoices:
    - DIRECTEUR: all
    - CLIENT: only their own
    """
    query = db.query(Invoice)

    if current_user.role != UserRole.DIRECTEUR:
        query = query.filter(Invoice.client_id == current_user.id)

    if status:
        query = query.filter(Invoice.status == status)

    return query.offset(skip).limit(limit).all()


@router.post("", response_model=InvoiceResponse, status_code=201)
def create_invoice(
    payload: InvoiceCreate,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db)
):
    """Create new invoice with line items (DIRECTEUR only)"""
    # Calculate totals
    subtotal = sum(item.quantity * item.unit_price for item in payload.items)
    tax_amount = subtotal * (payload.tax_rate / 100)
    total = subtotal + tax_amount

    service_start = payload.service_start_date
    duration_months = payload.duration_months
    service_end = None

    if payload.contract_id:
        contract = db.query(Contract).filter(Contract.id == payload.contract_id).first()
        if contract:
            service_start = service_start or contract.start_date
            duration_months = duration_months or contract.duration_months

    if service_start and duration_months:
        service_end = add_months(service_start, int(duration_months))

    invoice = Invoice(
        invoice_number=generate_invoice_number(db),
        client_id=payload.client_id,
        contract_id=payload.contract_id,
        service_start_date=service_start,
        service_end_date=service_end,
        duration_months=duration_months,
        due_date=payload.due_date,
        tax_rate=payload.tax_rate,
        subtotal=subtotal,
        tax_amount=tax_amount,
        total=total,
        notes=payload.notes,
        created_by=current_user.id
    )
    db.add(invoice)
    db.flush()

    # Add line items
    for item in payload.items:
        db_item = InvoiceItem(
            invoice_id=invoice.id,
            description=item.description,
            quantity=item.quantity,
            unit_price=item.unit_price,
            total=item.quantity * item.unit_price
        )
        db.add(db_item)

    db.commit()
    db.refresh(invoice)
    return invoice


@router.get("/export/excel")
def export_invoices_excel(
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db)
):
    """Export all invoices to Excel (DIRECTEUR only)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    invoices = db.query(Invoice).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a1a2e")

    headers = ["#", "Invoice No", "Client ID", "Status", "Issue Date",
               "Service Start", "Service End", "Duration", "Due Date", "Subtotal", "Tax %", "Tax Amount", "Total", "Currency"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, inv in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=inv.invoice_number)
        ws.cell(row=row, column=3, value=str(inv.client_id))
        ws.cell(row=row, column=4, value=inv.status.value)
        ws.cell(row=row, column=5, value=inv.issue_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=6, value=inv.service_start_date.strftime("%Y-%m-%d") if inv.service_start_date else "")
        ws.cell(row=row, column=7, value=inv.service_end_date.strftime("%Y-%m-%d") if inv.service_end_date else "")
        ws.cell(row=row, column=8, value=inv.duration_months or "")
        ws.cell(row=row, column=9, value=inv.due_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=10, value=inv.subtotal)
        ws.cell(row=row, column=11, value=inv.tax_rate)
        ws.cell(row=row, column=12, value=inv.tax_amount)
        ws.cell(row=row, column=13, value=inv.total)
        ws.cell(row=row, column=14, value=inv.currency)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices.xlsx"}
    )

@router.get("/export/accounting-canvas")
def export_accounting_canvas(
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db)
):
    """Preparation canva comptabilité vente client."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    invoices = db.query(Invoice).options(joinedload(Invoice.client).joinedload(User.profile), joinedload(Invoice.items)).order_by(Invoice.issue_date.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Canva Vente Client"

    ws.merge_cells("A1:K1")
    ws["A1"] = "UNIVERSAL INVEST STRATEGY - CANVA COMPTABILITE VENTE CLIENT"
    ws["A1"].font = Font(bold=True, size=14, color="173B57")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Date", "N° Facture", "Client", "ICE Client", "Période début", "Période fin", "Durée", "Désignation", "HT", "TVA %", "TVA", "TTC", "Statut", "Mode"]
    header_fill = PatternFill("solid", fgColor="EAF0F5")
    thin = Side(style="thin", color="CBD5E1")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="173B57")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    row = 4
    for inv in invoices:
        p = inv.client.profile if inv.client and inv.client.profile else None
        client = getattr(p, "company_name", None) or (f"{getattr(p, 'first_name', '')} {getattr(p, 'last_name', '')}".strip() if p else str(inv.client_id))
        ice = getattr(p, "company_ice", "") if p else ""
        designation = " | ".join([it.description for it in inv.items])
        values = [
            inv.issue_date.strftime("%d/%m/%Y"), inv.invoice_number, client, ice,
            inv.service_start_date.strftime("%d/%m/%Y") if inv.service_start_date else "",
            inv.service_end_date.strftime("%d/%m/%Y") if inv.service_end_date else "",
            f"{inv.duration_months} mois" if inv.duration_months else "",
            designation, inv.subtotal, inv.tax_rate, inv.tax_amount, inv.total,
            inv.status.value, "Vente client"
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        row += 1

    widths = [13, 18, 32, 22, 14, 14, 12, 50, 14, 10, 14, 14, 14, 18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=canva-comptabilite-vente-client.xlsx"}
    )



def _sage_amount(value) -> str:
    return f"{float(value or 0):.2f}".replace(".", ",")


def _sage_date(dt) -> str:
    return dt.strftime("%Y%m%d") if dt else ""


def _sage_clean(value) -> str:
    return str(value or "").replace(";", ",").replace("\r", " ").replace("\n", " ").strip()


def _invoice_designation(inv: Invoice) -> str:
    if inv.items:
        return " | ".join([_sage_clean(it.description) for it in inv.items if it.description]) or "Vente prestation"
    return "Vente prestation"


@router.get("/export/sage-txt")
def export_sage_txt(
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db)
):
    """Export Sage 100 Comptabilité i7 compatible TXT.

    Format minimal:
    Journal;Date;Piece;Compte;Libelle;Debit;Credit

    Pour éviter l'erreur Sage "Format de fichier incorrect":
    - pas d'entête/header;
    - date en YYYYMMDD;
    - montants avec virgule décimale;
    - encodage ANSI / Windows-1252;
    - fins de ligne CRLF Windows.
    """
    invoices = (
        db.query(Invoice)
        .options(joinedload(Invoice.client).joinedload(User.profile), joinedload(Invoice.items))
        .order_by(Invoice.issue_date.asc())
        .all()
    )

    lines = []
    for inv in invoices:
        p = inv.client.profile if inv.client and inv.client.profile else None
        client_name = _sage_clean(
            getattr(p, "company_name", None)
            or f"{getattr(p, 'first_name', '')} {getattr(p, 'last_name', '')}".strip()
            or getattr(inv.client, "email", None)
            or inv.client_id
        )
        date = _sage_date(inv.issue_date)
        piece = _sage_clean(inv.invoice_number)
        designation = _sage_clean(_invoice_designation(inv))
        tax_label = _sage_clean(f"TVA {inv.tax_rate:g}% - {designation}")

        # 3421 = Clients, 712 = Ventes/prestations, 4455 = TVA facturée.
        lines.append(";".join(["VTE", date, piece, "3421", f"Client {client_name}", _sage_amount(inv.total), _sage_amount(0)]))
        lines.append(";".join(["VTE", date, piece, "712", designation, _sage_amount(0), _sage_amount(inv.subtotal)]))
        if float(inv.tax_amount or 0) != 0:
            lines.append(";".join(["VTE", date, piece, "4455", tax_label, _sage_amount(0), _sage_amount(inv.tax_amount)]))

    content = "\r\n".join(lines) + "\r\n"
    return StreamingResponse(
        io.BytesIO(content.encode("cp1252", errors="replace")),
        media_type="text/plain; charset=windows-1252",
        headers={"Content-Disposition": "attachment; filename=SAGE_VENTES_COMPATIBLE.txt"}
    )


@router.get("/export/sage-rich-txt")
def export_sage_rich_txt(
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db)
):
    """Export TXT enrichi pour archivage/interne, pas le format direct recommandé pour Sage."""
    invoices = (
        db.query(Invoice)
        .options(joinedload(Invoice.client).joinedload(User.profile), joinedload(Invoice.items))
        .order_by(Invoice.issue_date.asc())
        .all()
    )
    lines = ["Journal;Date;Piece;Compte;Client;ICE;Libelle;Debit;Credit;TVA;TTC;PeriodeDebut;PeriodeFin"]
    for inv in invoices:
        p = inv.client.profile if inv.client and inv.client.profile else None
        client = _sage_clean(getattr(p, "company_name", None) or f"{getattr(p, 'first_name', '')} {getattr(p, 'last_name', '')}".strip() or str(inv.client_id))
        ice = _sage_clean(getattr(p, "company_ice", "") if p else "")
        libelle = _sage_clean(_invoice_designation(inv))
        date = inv.issue_date.strftime("%d/%m/%Y")
        start = inv.service_start_date.strftime("%d/%m/%Y") if inv.service_start_date else ""
        end = inv.service_end_date.strftime("%d/%m/%Y") if inv.service_end_date else ""
        lines.append(f"VTE;{date};{inv.invoice_number};3421;{client};{ice};{libelle};{inv.total:.2f};0.00;{inv.tax_amount:.2f};{inv.total:.2f};{start};{end}")
        lines.append(f"VTE;{date};{inv.invoice_number};712;{client};{ice};{libelle};0.00;{inv.subtotal:.2f};{inv.tax_amount:.2f};{inv.total:.2f};{start};{end}")
        if inv.tax_amount:
            lines.append(f"VTE;{date};{inv.invoice_number};4455;{client};{ice};TVA {inv.tax_rate:g}% - {libelle};0.00;{inv.tax_amount:.2f};{inv.tax_amount:.2f};{inv.total:.2f};{start};{end}")
    content = "\n".join(lines) + "\n"
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=SAGE_VENTES_RICH_ARCHIVE.txt"}
    )




def _safe_invoice_filename(invoice: Invoice) -> str:
    raw = invoice.invoice_number or f"invoice-{invoice.id}"
    safe = "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in raw)
    return f"{safe}.pdf"


def _get_invoice_for_file(invoice_id: uuid.UUID, current_user: User, db: Session) -> Invoice:
    invoice = (
        db.query(Invoice)
        .options(
            joinedload(Invoice.client).joinedload(User.profile),
            joinedload(Invoice.items),
            joinedload(Invoice.contract),
        )
        .filter(Invoice.id == invoice_id)
        .first()
    )
    if not invoice:
        raise HTTPException(status_code=404, detail="Facture introuvable")
    check_resource_access(invoice.client_id, current_user)
    return invoice


def _build_invoice_pdf_response(invoice: Invoice) -> FileResponse:
    try:
        path = generate_invoice_pdf(invoice)
        if not path or not os.path.exists(path):
            raise RuntimeError("Le fichier PDF n'a pas été généré")
        absolute_path = str(Path(path).resolve())
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erreur génération PDF facture: {exc}")

    filename = _safe_invoice_filename(invoice)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Access-Control-Expose-Headers": "Content-Disposition",
        "Cache-Control": "no-store",
    }
    return FileResponse(absolute_path, media_type="application/pdf", filename=filename, headers=headers)


@router.get("/{invoice_id}/download/pdf")
def download_invoice_pdf(
    invoice_id: uuid.UUID,
    current_user: User = Depends(require_any_authenticated),
    db: Session = Depends(get_db)
):
    """Download invoice PDF. Directeur: all invoices. Client: own invoices only."""
    invoice = _get_invoice_for_file(invoice_id, current_user, db)
    return _build_invoice_pdf_response(invoice)


@router.get("/{invoice_id}/download")
def download_invoice_pdf_alias(
    invoice_id: uuid.UUID,
    current_user: User = Depends(require_any_authenticated),
    db: Session = Depends(get_db)
):
    """Backward-compatible alias for PDF download."""
    invoice = _get_invoice_for_file(invoice_id, current_user, db)
    return _build_invoice_pdf_response(invoice)


@router.post("/{invoice_id}/send-email")
def send_invoice_email(
    invoice_id: uuid.UUID,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db)
):
    """Resend a generated invoice PDF to the client/company email."""
    invoice = _get_invoice_for_file(invoice_id, current_user, db)
    recipient = get_client_email(invoice)
    if not recipient:
        raise HTTPException(status_code=400, detail="Email client introuvable. Ajoute company_email ou email utilisateur.")
    try:
        pdf_path = generate_invoice_pdf(invoice)
        if not pdf_path or not os.path.exists(pdf_path):
            raise RuntimeError("Le PDF de facture n'a pas été généré")
        sent = send_email_with_attachments(
            recipient,
            f"Votre facture {invoice.invoice_number} - Universal Invest Strategy",
            "Bonjour,\n\nVeuillez trouver votre facture en pièce jointe.\n\nCordialement,\nUniversal Invest Strategy",
            [pdf_path],
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Erreur envoi email facture: {exc}")
    if not sent:
        raise HTTPException(status_code=400, detail="SMTP non configuré. Vérifie SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD et SMTP_FROM_EMAIL dans backend/.env")
    return {"sent": True, "to": recipient, "invoice_number": invoice.invoice_number}


@router.get("/{invoice_id}", response_model=InvoiceResponse)
def get_invoice(
    invoice_id: uuid.UUID,
    current_user: User = Depends(require_any_authenticated),
    db: Session = Depends(get_db)
):
    """Get invoice by ID"""
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    check_resource_access(invoice.client_id, current_user)
    return invoice


@router.put("/{invoice_id}", response_model=InvoiceResponse)
def update_invoice(
    invoice_id: uuid.UUID,
    payload: InvoiceUpdate,
    current_user: User = Depends(require_directeur),
    db: Session = Depends(get_db)
):
    """Update invoice status/details (DIRECTEUR only)"""
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    update_data = payload.model_dump(exclude_unset=True)
    if "service_start_date" in update_data or "duration_months" in update_data:
        service_start = update_data.get("service_start_date") or invoice.service_start_date
        duration = update_data.get("duration_months") or invoice.duration_months
        if service_start and duration:
            update_data["service_end_date"] = add_months(service_start, int(duration))
    for field, value in update_data.items():
        setattr(invoice, field, value)

    if payload.status == InvoiceStatus.PAID and not invoice.paid_date:
        invoice.paid_date = datetime.utcnow()

    db.commit()
    db.refresh(invoice)
    return invoice
